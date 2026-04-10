#!/usr/bin/env python3
"""
全来店报价单 PDF + Excel 生成器
用法: python generate_quotation.py --config quotation_data.json --output 报价单.pdf [--output-xlsx 报价单.xlsx] [--profit]

输入: JSON 配置文件（包含客户信息、产品选择、折扣等）
输出: PDF 报价单 + 可选 Excel 报价单
可选: --profit 在终端输出利润测算（不写入文档）
"""

import json
import sys
import os
import argparse
from datetime import datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP

# ============================================================
# reportlab imports
# ============================================================
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm, cm
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    PageBreak, KeepTogether
)
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# ============================================================
# 注册中文字体
# ============================================================
# 策略：用 Helvetica 处理数字/英文，DroidSansFallbackFull 处理中文
# 在 Paragraph 中通过 <font> 标签混排

_CJK_FONT_NAME = 'CJKFont'
_LATIN_FONT_NAME = 'Helvetica'
_CN_FONT_NAME = _CJK_FONT_NAME  # 默认字体名（用于 TableStyle 等）

# 查找并注册 CJK 字体
_FONT_CANDIDATES = [
    '/usr/share/fonts/truetype/droid/DroidSansFallbackFull.ttf',
    '/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc',
    '/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc',
    '/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc',
    '/usr/share/fonts/truetype/wqy/wqy-microhei.ttc',
]
# 也检查 Skill assets 目录中的自带字体
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_ASSETS_DIR = os.path.join(os.path.dirname(_SCRIPT_DIR), 'assets')
_FONT_CANDIDATES.insert(0, os.path.join(_ASSETS_DIR, 'NotoSansSC-Regular.ttf'))
_FONT_CANDIDATES.insert(0, os.path.join(_ASSETS_DIR, 'NotoSansSC-Regular.otf'))

_cjk_font_path = None
_is_complete_font = False  # 是否包含 Latin 字符

for _fp in _FONT_CANDIDATES:
    if os.path.exists(_fp):
        try:
            pdfmetrics.registerFont(TTFont(_CJK_FONT_NAME, _fp))
            _cjk_font_path = _fp
            # 检查是否包含数字（完整字体）
            from fontTools.ttLib import TTFont as FTFont
            _ft = FTFont(_fp)
            _cmap = _ft.getBestCmap()
            _is_complete_font = all(ord(c) in _cmap for c in '0123456789')
            _ft.close()
            break
        except Exception:
            continue

if _cjk_font_path is None:
    # 回退到 CID 字体
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    pdfmetrics.registerFont(UnicodeCIDFont('STSong-Light'))
    _CJK_FONT_NAME = 'STSong-Light'
    _CN_FONT_NAME = 'STSong-Light'
    _is_complete_font = True  # CID 字体通常包含 Latin

if _is_complete_font:
    # 完整字体，直接用一个字体名
    _CN_FONT_NAME = _CJK_FONT_NAME
else:
    # 不完整字体：以 CJK 字体为 Paragraph 基础字体
    # 将 ASCII/Latin 字符用 Helvetica <font> 标签包裹
    _CN_FONT_NAME = _CJK_FONT_NAME

import re

def _mixed_text(text):
    """
    翻转策略：以 CJK 字体作为 Paragraph 默认字体，
    将连续的 ASCII/Latin 字符用 <font name="Helvetica"> 包裹。
    这样所有中文字符（包括标点）自然用 CJK 字体渲染。
    """
    if _is_complete_font:
        return str(text)
    text = str(text)
    if not text:
        return text
    # 匹配连续的 ASCII 可打印字符（数字、英文、标点、空格等）
    # ASCII 范围: 0x0020-0x007E
    ascii_pattern = re.compile(r'([\x20-\x7e]+)')
    result = ascii_pattern.sub(
        lambda m: f'<font name="{_LATIN_FONT_NAME}">{m.group(0)}</font>',
        text
    )
    return result

# ============================================================
# 颜色定义
# ============================================================
HEADER_BG = colors.HexColor('#FFB300')      # 金黄表头
HEADER_FG = colors.white
ROW_ALT_BG = colors.HexColor('#FFFBF0')     # 交替行背景
BORDER_COLOR = colors.HexColor('#d0d5dd')
TOTAL_BG = colors.HexColor('#FFF5D6')       # 合计行背景
ACCENT = colors.HexColor('#FFB300')          # 强调色

# ============================================================
# 样式定义
# ============================================================
def get_styles():
    styles = getSampleStyleSheet()

    styles.add(ParagraphStyle(
        name='CNTitle',
        fontName=_CN_FONT_NAME,
        fontSize=18,
        leading=24,
        alignment=1,  # center
        spaceAfter=6,
    ))
    styles.add(ParagraphStyle(
        name='CNSubtitle',
        fontName=_CN_FONT_NAME,
        fontSize=11,
        leading=15,
        alignment=1,
        textColor=colors.HexColor('#555555'),
        spaceAfter=12,
    ))
    styles.add(ParagraphStyle(
        name='CNNormal',
        fontName=_CN_FONT_NAME,
        fontSize=9,
        leading=13,
    ))
    styles.add(ParagraphStyle(
        name='CNSmall',
        fontName=_CN_FONT_NAME,
        fontSize=8,
        leading=11,
    ))
    styles.add(ParagraphStyle(
        name='CNBold',
        fontName=_CN_FONT_NAME,
        fontSize=9,
        leading=13,
    ))
    styles.add(ParagraphStyle(
        name='CNSection',
        fontName=_CN_FONT_NAME,
        fontSize=12,
        leading=16,
        spaceBefore=12,
        spaceAfter=6,
    ))
    styles.add(ParagraphStyle(
        name='CNFooter',
        fontName=_CN_FONT_NAME,
        fontSize=8,
        leading=11,
        textColor=colors.HexColor('#888888'),
    ))
    styles.add(ParagraphStyle(
        name='CellStyle',
        fontName=_CN_FONT_NAME,
        fontSize=8,
        leading=11,
        wordWrap='CJK',
    ))
    styles.add(ParagraphStyle(
        name='CellStyleRight',
        fontName=_CN_FONT_NAME,
        fontSize=8,
        leading=11,
        alignment=2,  # right
    ))
    styles.add(ParagraphStyle(
        name='CellStyleCenter',
        fontName=_CN_FONT_NAME,
        fontSize=8,
        leading=11,
        alignment=1,  # center
    ))
    return styles

# ============================================================
# 辅助函数
# ============================================================
def fmt_money(val):
    """格式化金额：带千分位逗号"""
    if val is None or val == '赠送':
        return '赠送'
    try:
        d = Decimal(str(val)).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)
        return '{:,.2f}'.format(float(d))
    except:
        return str(val)

def fmt_pct(val):
    """格式化折扣百分比"""
    if val is None or val == 0:
        return '-'
    return f'{val*100:.0f}%'

def number_to_chinese(num):
    """数字转大写中文金额"""
    chinese_digits = '零壹贰叁肆伍陆柒捌玖'
    chinese_units = ['', '拾', '佰', '仟']
    chinese_big_units = ['', '万', '亿']

    if num == 0:
        return '零元整'

    d = Decimal(str(num)).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)
    int_part = int(d)
    dec_part = int(round((float(d) - int_part) * 100))

    result = ''
    if int_part > 0:
        s = str(int_part)
        n = len(s)
        for i, ch in enumerate(s):
            digit = int(ch)
            pos = n - 1 - i
            big_unit_idx = pos // 4
            unit_idx = pos % 4

            if digit != 0:
                result += chinese_digits[digit] + chinese_units[unit_idx]
                if unit_idx == 0 and big_unit_idx > 0:
                    result += chinese_big_units[big_unit_idx]
            else:
                if unit_idx == 0 and big_unit_idx > 0:
                    result += chinese_big_units[big_unit_idx]
                elif result and not result.endswith('零'):
                    result += '零'

        result = result.rstrip('零')
        result += '元'

    if dec_part == 0:
        result += '整'
    else:
        jiao = dec_part // 10
        fen = dec_part % 10
        if jiao > 0:
            result += chinese_digits[jiao] + '角'
        if fen > 0:
            result += chinese_digits[fen] + '分'

    return result

def gen_quote_number():
    """生成报价编号: QLT-YYYYMMDD-XXX"""
    now = datetime.now()
    return f"QLT-{now.strftime('%Y%m%d')}-001"

# ============================================================
# 标准模板（≤50店）— PDF
# ============================================================
def build_standard_template(data, styles):
    """生成标准单页报价单"""
    story = []

    # === 页眉/标题 ===
    story.append(Paragraph(_mixed_text('"全来店"产品报价单'), styles['CNTitle']))
    story.append(Paragraph(
        _mixed_text('上海收钱吧互联网科技股份有限公司'),
        styles['CNSubtitle']
    ))
    story.append(Spacer(1, 4*mm))

    # === 客户信息 ===
    client = data.get('客户信息', {})
    quote_no = data.get('报价编号', gen_quote_number())
    quote_date = data.get('报价日期', datetime.now().strftime('%Y年%m月%d日'))
    validity = data.get('报价有效期', '30个工作日')

    info_data = [
        [Paragraph(_mixed_text(f'致：{client.get("公司名称", "")}'), styles['CNNormal']),
         Paragraph(_mixed_text(f'报价编号：{quote_no}'), styles['CNNormal'])],
        [Paragraph(_mixed_text(f'联系人：{client.get("联系人", "")}'), styles['CNNormal']),
         Paragraph(_mixed_text(f'报价日期：{quote_date}'), styles['CNNormal'])],
        [Paragraph(_mixed_text(f'地址：{client.get("地址", "")}'), styles['CNNormal']),
         Paragraph(_mixed_text(f'有效期：{validity}'), styles['CNNormal'])],
    ]
    if client.get('电话'):
        info_data.append([
            Paragraph(_mixed_text(f'电话：{client.get("电话", "")}'), styles['CNNormal']),
            Paragraph(_mixed_text(''), styles['CNNormal'])
        ])

    info_table = Table(info_data, colWidths=[95*mm, 75*mm])
    info_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('TOPPADDING', (0,0), (-1,-1), 2),
        ('BOTTOMPADDING', (0,0), (-1,-1), 2),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 6*mm))

    # === 报价明细表 ===
    items = data.get('报价项目', [])

    # 表头
    col_widths = [10*mm, 28*mm, 38*mm, 16*mm, 20*mm, 14*mm, 16*mm, 20*mm, 22*mm]
    header = [
        Paragraph(_mixed_text('序号'), styles['CellStyleCenter']),
        Paragraph(_mixed_text('商品分类'), styles['CellStyleCenter']),
        Paragraph(_mixed_text('商品名称'), styles['CellStyleCenter']),
        Paragraph(_mixed_text('单位'), styles['CellStyleCenter']),
        Paragraph(_mixed_text('标准价'), styles['CellStyleCenter']),
        Paragraph(_mixed_text('折扣'), styles['CellStyleCenter']),
        Paragraph(_mixed_text('数量'), styles['CellStyleCenter']),
        Paragraph(_mixed_text('实际价'), styles['CellStyleCenter']),
        Paragraph(_mixed_text('小计'), styles['CellStyleCenter']),
    ]

    table_data = [header]
    total = Decimal('0')

    for idx, item in enumerate(items, 1):
        std_price = item.get('标准价', 0)
        discount = item.get('折扣', 0)
        qty = item.get('数量', 1)

        if std_price == '赠送' or std_price is None:
            actual_price = '赠送'
            subtotal = '赠送'
        else:
            std_price_d = Decimal(str(std_price))
            discount_d = Decimal(str(discount))
            qty_d = Decimal(str(qty))
            actual_price_d = (std_price_d * (1 - discount_d)).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)
            subtotal_d = (actual_price_d * qty_d).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)
            actual_price = float(actual_price_d)
            subtotal = float(subtotal_d)
            total += subtotal_d

        row = [
            Paragraph(_mixed_text(idx), styles['CellStyleCenter']),
            Paragraph(_mixed_text(item.get('商品分类', '')), styles['CellStyle']),
            Paragraph(_mixed_text(item.get('商品名称', '')), styles['CellStyle']),
            Paragraph(_mixed_text(item.get('单位', '')), styles['CellStyleCenter']),
            Paragraph(_mixed_text(fmt_money(std_price)), styles['CellStyleRight']),
            Paragraph(_mixed_text(fmt_pct(discount)), styles['CellStyleCenter']),
            Paragraph(_mixed_text(qty), styles['CellStyleCenter']),
            Paragraph(_mixed_text(fmt_money(actual_price)), styles['CellStyleRight']),
            Paragraph(_mixed_text(fmt_money(subtotal)), styles['CellStyleRight']),
        ]
        table_data.append(row)

    # 合计行
    total_float = float(total)
    total_row = [
        Paragraph(_mixed_text(''), styles['CellStyle']),
        Paragraph(_mixed_text(''), styles['CellStyle']),
        Paragraph(_mixed_text(''), styles['CellStyle']),
        Paragraph(_mixed_text(''), styles['CellStyle']),
        Paragraph(_mixed_text(''), styles['CellStyle']),
        Paragraph(_mixed_text(''), styles['CellStyle']),
        Paragraph(_mixed_text('合计'), styles['CellStyleCenter']),
        Paragraph(_mixed_text(''), styles['CellStyle']),
        Paragraph(_mixed_text(fmt_money(total_float)), styles['CellStyleRight']),
    ]
    table_data.append(total_row)

    # 构建表格
    t = Table(table_data, colWidths=col_widths, repeatRows=1)

    style_cmds = [
        # 表头样式
        ('BACKGROUND', (0,0), (-1,0), HEADER_BG),
        ('TEXTCOLOR', (0,0), (-1,0), HEADER_FG),
        # 全局
        ('FONTNAME', (0,0), (-1,-1), _CN_FONT_NAME),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('LEFTPADDING', (0,0), (-1,-1), 3),
        ('RIGHTPADDING', (0,0), (-1,-1), 3),
        # 网格
        ('GRID', (0,0), (-1,-1), 0.5, BORDER_COLOR),
        # 合计行样式
        ('BACKGROUND', (0,-1), (-1,-1), TOTAL_BG),
        ('SPAN', (0,-1), (5,-1)),
    ]

    # 交替行颜色
    for i in range(1, len(table_data) - 1):
        if i % 2 == 0:
            style_cmds.append(('BACKGROUND', (0,i), (-1,i), ROW_ALT_BG))

    t.setStyle(TableStyle(style_cmds))
    story.append(t)
    story.append(Spacer(1, 4*mm))

    # === 金额大写 ===
    chinese_total = number_to_chinese(total_float)
    story.append(Paragraph(
        _mixed_text(f'合计金额（大写）：{chinese_total}'),
        styles['CNNormal']
    ))
    story.append(Spacer(1, 8*mm))

    # === 备注条款 ===
    story.append(Paragraph(_mixed_text('备注：'), styles['CNSection']))
    terms = data.get('条款', [
        '以上报价金额均为含税金额，税率为6%；',
        '报价有效期为30个工作日，自报价单生成之日起；',
        '具体折扣金额按签订合同（或销售订单）时具体数量确定价格；',
        '涉及短信、小程序授权、外卖平台接口调用等第三方机构收费部分，需单独计费；',
        '如需要三方代仓对接，需要一事一议。',
    ])
    for i, term in enumerate(terms, 1):
        cn_num = '①②③④⑤⑥⑦⑧⑨⑩'[i-1] if i <= 10 else f'{i}.'
        story.append(Paragraph(_mixed_text(f'{cn_num} {term}'), styles['CNSmall']))

    story.append(Spacer(1, 12*mm))

    # === 页脚公司信息 ===
    story.append(Paragraph(
        _mixed_text('上海收钱吧互联网科技股份有限公司'),
        styles['CNNormal']
    ))
    story.append(Paragraph(
        _mixed_text('地址：上海市闵行区浦江智慧广场陈行公路2168号7号楼'),
        styles['CNFooter']
    ))

    return story

# ============================================================
# 阶梯报价对比页（定制版附页）
# ============================================================
def _build_tiered_section(data, styles):
    """构建阶梯报价对比页，按门店规模展示不同折扣下的费用对比"""
    tiers = data.get('阶梯配置', [])
    if not tiers:
        return []

    items = data.get('报价项目', [])
    n_tiers = len(tiers)

    story = [PageBreak()]
    story.append(Paragraph(_mixed_text('阶梯报价参考'), styles['CNSection']))
    story.append(Spacer(1, 2*mm))
    story.append(Paragraph(
        _mixed_text('以下为不同门店规模下的费用参考，实际以签订合同为准。'),
        styles['CNSmall']
    ))
    story.append(Spacer(1, 3*mm))

    # 列宽：商品名称42 | 单位12 | 标准价20 | 各tier平分剩余
    remaining = 180 - 42 - 12 - 20
    tier_col_w = remaining * mm / n_tiers
    col_widths = [42*mm, 12*mm, 20*mm] + [tier_col_w] * n_tiers

    # 表头行
    def tier_label(t):
        d = t.get('折扣', 0)
        if d == 0:
            return f"{t['标签']}\n（标准价）"
        zhe = round((1 - d) * 10, 1)
        zhe_str = f"{int(zhe)}折" if zhe == int(zhe) else f"{zhe}折"
        return f"{t['标签']}\n（{zhe_str}）"

    header = [
        Paragraph(_mixed_text('商品名称'), styles['CellStyleCenter']),
        Paragraph(_mixed_text('单位'), styles['CellStyleCenter']),
        Paragraph(_mixed_text('标准价'), styles['CellStyleCenter']),
    ] + [Paragraph(_mixed_text(tier_label(t)), styles['CellStyleCenter']) for t in tiers]

    table_data = [header]

    # 按模块分组（排除硬件设备）
    cat_order = ['门店软件套餐', '门店增值模块', '总部模块', '实施服务']
    categories = {k: [] for k in cat_order}
    for item in items:
        cat = item.get('模块分类', '门店软件套餐')
        if cat == '硬件设备':
            continue
        if cat in categories:
            categories[cat].append(item)
        else:
            categories['门店软件套餐'].append(item)

    tier_grand_totals = [Decimal('0')] * n_tiers
    cat_header_rows = []   # 记录分类标题行索引
    subtotal_rows = []     # 记录小计行索引

    # 实施服务不受阶梯折扣影响
    NO_TIER_DISCOUNT_CATS = {'实施服务'}

    for cat_name in cat_order:
        cat_items = categories[cat_name]
        if not cat_items:
            continue

        cat_row_idx = len(table_data)
        cat_header_rows.append(cat_row_idx)
        cat_row = [Paragraph(_mixed_text(cat_name), styles['CNBold'])] + \
                  [Paragraph('', styles['CellStyle'])] * (2 + n_tiers)
        table_data.append(cat_row)

        cat_tier_totals = [Decimal('0')] * n_tiers
        apply_tier_discount = cat_name not in NO_TIER_DISCOUNT_CATS

        for item in cat_items:
            std_price = item.get('标准价', 0)
            unit = item.get('单位', '')
            item_qty = item.get('数量', 1)
            is_per_store = '店' in unit

            if std_price == '赠送' or std_price is None:
                row = [
                    Paragraph(_mixed_text(item.get('商品名称', '')), styles['CellStyle']),
                    Paragraph(_mixed_text(unit), styles['CellStyleCenter']),
                    Paragraph(_mixed_text('赠送'), styles['CellStyleCenter']),
                ] + [Paragraph(_mixed_text('赠送'), styles['CellStyleCenter']) for _ in tiers]
            else:
                std_d = Decimal(str(std_price))
                row = [
                    Paragraph(_mixed_text(item.get('商品名称', '')), styles['CellStyle']),
                    Paragraph(_mixed_text(unit), styles['CellStyleCenter']),
                    Paragraph(_mixed_text(fmt_money(float(std_d))), styles['CellStyleRight']),
                ]
                for ti, t in enumerate(tiers):
                    d = Decimal(str(t.get('折扣', 0))) if apply_tier_discount else Decimal('0')
                    qty = Decimal(str(t['门店数'])) if is_per_store else Decimal(str(item_qty))
                    actual = (std_d * (1 - d)).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)
                    subtotal = (actual * qty).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)
                    cat_tier_totals[ti] += subtotal
                    row.append(Paragraph(_mixed_text(fmt_money(float(subtotal))), styles['CellStyleRight']))
            table_data.append(row)

        # 分类小计行
        sub_row_idx = len(table_data)
        subtotal_rows.append(sub_row_idx)
        sub_row = [
            Paragraph('', styles['CellStyle']),
            Paragraph('', styles['CellStyle']),
            Paragraph(_mixed_text('小计'), styles['CellStyleCenter']),
        ]
        for ti, tot in enumerate(cat_tier_totals):
            tier_grand_totals[ti] += tot
            sub_row.append(Paragraph(_mixed_text(fmt_money(float(tot))), styles['CellStyleRight']))
        table_data.append(sub_row)

    # 合计行
    total_row = [
        Paragraph('', styles['CellStyle']),
        Paragraph('', styles['CellStyle']),
        Paragraph(_mixed_text('合计'), styles['CNBold']),
    ] + [Paragraph(_mixed_text(f'¥ {fmt_money(float(tot))}'), styles['CellStyleRight'])
         for tot in tier_grand_totals]
    table_data.append(total_row)

    # 折算单店年费行
    unit_row = [
        Paragraph(_mixed_text('折算单店年费'), styles['CellStyle']),
        Paragraph('', styles['CellStyle']),
        Paragraph('', styles['CellStyle']),
    ]
    for ti, t in enumerate(tiers):
        stores = Decimal(str(t['门店数']))
        per_store = (tier_grand_totals[ti] / stores).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)
        unit_row.append(Paragraph(_mixed_text(fmt_money(float(per_store))), styles['CellStyleRight']))
    table_data.append(unit_row)

    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    n_rows = len(table_data)
    style_cmds = [
        ('BACKGROUND', (0, 0), (-1, 0), HEADER_BG),
        ('TEXTCOLOR', (0, 0), (-1, 0), HEADER_FG),
        ('FONTNAME', (0, 0), (-1, -1), _CN_FONT_NAME),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('GRID', (0, 0), (-1, -1), 0.5, BORDER_COLOR),
        ('BACKGROUND', (0, -1), (-1, -1), TOTAL_BG),   # 单店年费行
        ('BACKGROUND', (0, -2), (-1, -2), TOTAL_BG),   # 合计行
    ]
    for row_idx in cat_header_rows:
        style_cmds += [
            ('SPAN', (0, row_idx), (-1, row_idx)),
            ('BACKGROUND', (0, row_idx), (-1, row_idx), ROW_ALT_BG),
        ]
    for row_idx in subtotal_rows:
        style_cmds.append(('BACKGROUND', (0, row_idx), (-1, row_idx), TOTAL_BG))

    t.setStyle(TableStyle(style_cmds))
    story.append(t)
    return story


# ============================================================
# 定制多页模板（>50店）— PDF
# ============================================================
def build_custom_template(data, styles):
    """生成定制多页报价单"""
    story = []

    # === 第1页：封面 ===
    story.append(Spacer(1, 15*mm))
    story.append(Paragraph(_mixed_text('"全来店"产品报价方案'), styles['CNTitle']))
    story.append(Spacer(1, 4*mm))
    story.append(Paragraph(
        _mixed_text('上海收钱吧互联网科技股份有限公司'),
        styles['CNSubtitle']
    ))
    story.append(Spacer(1, 10*mm))

    client = data.get('客户信息', {})
    quote_no = data.get('报价编号', gen_quote_number())
    quote_date = data.get('报价日期', datetime.now().strftime('%Y年%m月%d日'))

    cover_info = [
        ['客户名称', client.get('公司名称', '')],
        ['联系人', client.get('联系人', '')],
        ['联系电话', client.get('电话', '')],
        ['报价编号', quote_no],
        ['报价日期', quote_date],
        ['有效期', data.get('报价有效期', '30个工作日')],
    ]

    cover_data = [[Paragraph(_mixed_text(r[0]), styles['CNBold']),
                    Paragraph(_mixed_text(r[1]), styles['CNNormal'])] for r in cover_info]

    cover_table = Table(cover_data, colWidths=[40*mm, 100*mm])
    cover_table.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), _CN_FONT_NAME),
        ('GRID', (0,0), (-1,-1), 0.5, BORDER_COLOR),
        ('BACKGROUND', (0,0), (0,-1), ROW_ALT_BG),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('LEFTPADDING', (0,0), (-1,-1), 6),
    ]))
    story.append(cover_table)
    story.append(Spacer(1, 8*mm))

    # === 分类汇总各模块 ===
    items = data.get('报价项目', [])

    # 按类别分组（定制版不含硬件设备）
    categories = {
        '门店软件套餐': [],
        '门店增值模块': [],
        '总部模块': [],
        '实施服务': [],
    }

    for item in items:
        cat = item.get('模块分类', '门店软件套餐')
        if cat in categories:
            categories[cat].append(item)
        elif cat != '硬件设备':
            categories['门店软件套餐'].append(item)

    grand_total = Decimal('0')
    first_cat = True

    for cat_name, cat_items in categories.items():
        if not cat_items:
            continue

        if first_cat:
            story.append(PageBreak())
            first_cat = False
        else:
            story.append(Spacer(1, 8*mm))
        story.append(Paragraph(_mixed_text(cat_name), styles['CNSection']))
        story.append(Spacer(1, 3*mm))

        col_widths = [10*mm, 28*mm, 40*mm, 16*mm, 20*mm, 14*mm, 16*mm, 20*mm, 22*mm]
        header = [
            Paragraph(_mixed_text('序号'), styles['CellStyleCenter']),
            Paragraph(_mixed_text('商品分类'), styles['CellStyleCenter']),
            Paragraph(_mixed_text('商品名称'), styles['CellStyleCenter']),
            Paragraph(_mixed_text('单位'), styles['CellStyleCenter']),
            Paragraph(_mixed_text('标准价'), styles['CellStyleCenter']),
            Paragraph(_mixed_text('折扣'), styles['CellStyleCenter']),
            Paragraph(_mixed_text('数量'), styles['CellStyleCenter']),
            Paragraph(_mixed_text('实际价'), styles['CellStyleCenter']),
            Paragraph(_mixed_text('小计'), styles['CellStyleCenter']),
        ]

        table_data = [header]
        cat_total = Decimal('0')

        for idx, item in enumerate(cat_items, 1):
            std_price = item.get('标准价', 0)
            discount = item.get('折扣', 0)
            qty = item.get('数量', 1)

            if std_price == '赠送' or std_price is None:
                actual_price = '赠送'
                subtotal = '赠送'
            else:
                std_price_d = Decimal(str(std_price))
                discount_d = Decimal(str(discount))
                qty_d = Decimal(str(qty))
                actual_price_d = (std_price_d * (1 - discount_d)).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)
                subtotal_d = (actual_price_d * qty_d).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)
                actual_price = float(actual_price_d)
                subtotal = float(subtotal_d)
                cat_total += subtotal_d

            row = [
                Paragraph(_mixed_text(idx), styles['CellStyleCenter']),
                Paragraph(_mixed_text(item.get('商品分类', '')), styles['CellStyle']),
                Paragraph(_mixed_text(item.get('商品名称', '')), styles['CellStyle']),
                Paragraph(_mixed_text(item.get('单位', '')), styles['CellStyleCenter']),
                Paragraph(_mixed_text(fmt_money(std_price)), styles['CellStyleRight']),
                Paragraph(_mixed_text(fmt_pct(discount)), styles['CellStyleCenter']),
                Paragraph(_mixed_text(qty), styles['CellStyleCenter']),
                Paragraph(_mixed_text(fmt_money(actual_price)), styles['CellStyleRight']),
                Paragraph(_mixed_text(fmt_money(subtotal)), styles['CellStyleRight']),
            ]
            table_data.append(row)

        # 小计行
        cat_total_float = float(cat_total)
        grand_total += cat_total

        subtotal_row = [
            Paragraph(_mixed_text(''), styles['CellStyle']),
            Paragraph(_mixed_text(''), styles['CellStyle']),
            Paragraph(_mixed_text(''), styles['CellStyle']),
            Paragraph(_mixed_text(''), styles['CellStyle']),
            Paragraph(_mixed_text(''), styles['CellStyle']),
            Paragraph(_mixed_text(''), styles['CellStyle']),
            Paragraph(_mixed_text('小计'), styles['CellStyleCenter']),
            Paragraph(_mixed_text(''), styles['CellStyle']),
            Paragraph(_mixed_text(fmt_money(cat_total_float)), styles['CellStyleRight']),
        ]
        table_data.append(subtotal_row)

        t = Table(table_data, colWidths=col_widths, repeatRows=1)
        style_cmds = [
            ('BACKGROUND', (0,0), (-1,0), HEADER_BG),
            ('TEXTCOLOR', (0,0), (-1,0), HEADER_FG),
            ('FONTNAME', (0,0), (-1,-1), _CN_FONT_NAME),
            ('FONTSIZE', (0,0), (-1,-1), 8),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ('LEFTPADDING', (0,0), (-1,-1), 3),
            ('RIGHTPADDING', (0,0), (-1,-1), 3),
            ('GRID', (0,0), (-1,-1), 0.5, BORDER_COLOR),
            ('BACKGROUND', (0,-1), (-1,-1), TOTAL_BG),
            ('SPAN', (0,-1), (5,-1)),
        ]
        for i in range(1, len(table_data) - 1):
            if i % 2 == 0:
                style_cmds.append(('BACKGROUND', (0,i), (-1,i), ROW_ALT_BG))
        t.setStyle(TableStyle(style_cmds))
        story.append(t)

    # === 最后一页：总计 + 条款 ===
    story.append(PageBreak())
    story.append(Paragraph(_mixed_text('费用汇总'), styles['CNSection']))
    story.append(Spacer(1, 4*mm))

    grand_total_float = float(grand_total)
    chinese_total = number_to_chinese(grand_total_float)

    summary_data = [
        [Paragraph(_mixed_text('项目总计（含税）'), styles['CNBold']),
         Paragraph(_mixed_text(f'¥ {fmt_money(grand_total_float)}'), styles['CellStyleRight'])],
        [Paragraph(_mixed_text('大写金额'), styles['CNBold']),
         Paragraph(_mixed_text(chinese_total), styles['CNNormal'])],
    ]
    summary_table = Table(summary_data, colWidths=[50*mm, 100*mm])
    summary_table.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), _CN_FONT_NAME),
        ('GRID', (0,0), (-1,-1), 0.5, BORDER_COLOR),
        ('BACKGROUND', (0,0), (0,-1), ROW_ALT_BG),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('LEFTPADDING', (0,0), (-1,-1), 6),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 6*mm))

    # 条款
    story.append(Paragraph(_mixed_text('备注与条款：'), styles['CNSection']))
    terms = data.get('条款', [
        '以上报价金额均为含税金额，税率为6%；',
        '报价有效期为30个工作日，自报价单生成之日起；',
        '具体折扣金额按签订合同（或销售订单）时具体数量确定价格；',
        '涉及短信、小程序授权、外卖平台接口调用等第三方机构收费部分，需单独计费；',
        '如需要三方代仓对接，需要一事一议。',
    ])
    for i, term in enumerate(terms, 1):
        cn_num = '①②③④⑤⑥⑦⑧⑨⑩'[i-1] if i <= 10 else f'{i}.'
        story.append(Paragraph(_mixed_text(f'{cn_num} {term}'), styles['CNSmall']))

    story.append(Spacer(1, 8*mm))
    story.append(Paragraph(
        _mixed_text('上海收钱吧互联网科技股份有限公司'),
        styles['CNNormal']
    ))
    story.append(Paragraph(
        _mixed_text('地址：上海市闵行区浦江智慧广场陈行公路2168号7号楼'),
        styles['CNFooter']
    ))

    # 阶梯报价对比页（如有配置）
    story.extend(_build_tiered_section(data, styles))

    return story

# ============================================================
# Excel 生成辅助函数
# ============================================================
def _xl_header_style(cell):
    """Excel 表头单元格样式"""
    from openpyxl.styles import Font, PatternFill, Alignment
    cell.font = Font(name='微软雅黑', bold=True, color='FFFFFF', size=10)
    cell.fill = PatternFill('solid', fgColor='FFB300')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def _xl_title_style(cell, size=16):
    """Excel 标题单元格样式"""
    from openpyxl.styles import Font, PatternFill, Alignment
    cell.font = Font(name='微软雅黑', bold=True, size=size, color='CC8800')
    cell.alignment = Alignment(horizontal='center', vertical='center')

def _xl_subtitle_style(cell):
    """Excel 副标题单元格样式"""
    from openpyxl.styles import Font, Alignment
    cell.font = Font(name='微软雅黑', size=11, color='555555')
    cell.alignment = Alignment(horizontal='center', vertical='center')

def _xl_info_label_style(cell):
    """Excel 客户信息标签样式"""
    from openpyxl.styles import Font, PatternFill, Alignment
    cell.font = Font(name='微软雅黑', bold=True, size=10, color='CC8800')
    cell.fill = PatternFill('solid', fgColor='FFFBF0')
    cell.alignment = Alignment(horizontal='left', vertical='center')

def _xl_info_value_style(cell):
    """Excel 客户信息值样式"""
    from openpyxl.styles import Font, Alignment
    cell.font = Font(name='微软雅黑', size=10)
    cell.alignment = Alignment(horizontal='left', vertical='center')

def _xl_data_style(cell, align='center', bold=False, bg=None, num_format=None):
    """Excel 数据单元格样式"""
    from openpyxl.styles import Font, PatternFill, Alignment
    cell.font = Font(name='微软雅黑', bold=bold, size=10)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    if bg:
        cell.fill = PatternFill('solid', fgColor=bg)
    if num_format:
        cell.number_format = num_format

def _xl_total_style(cell, align='right'):
    """Excel 合计行样式"""
    from openpyxl.styles import Font, PatternFill, Alignment
    cell.font = Font(name='微软雅黑', bold=True, size=10)
    cell.fill = PatternFill('solid', fgColor='FFF5D6')
    cell.alignment = Alignment(horizontal=align, vertical='center')

def _xl_apply_border(ws, min_row, min_col, max_row, max_col):
    """为 Excel 单元格区域应用细边框"""
    from openpyxl.styles import Border, Side
    thin = Side(style='thin', color='D0D5DD')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=min_row, min_col=min_col,
                             max_row=max_row, max_col=max_col):
        for cell in row:
            cell.border = border

def _xl_write_item_table(ws, items, start_row, sheet_name='', compute_values=False):
    """
    向 worksheet 写入报价明细表。
    列：A=序号, B=商品分类, C=商品名称, D=单位, E=标准价,
        F=折扣, G=数量, H=实际价, I=小计
    compute_values=True 时直接写计算后的数值（兼容性更好），
    否则写 Excel 公式（便于手动调整）。
    返回: (最后数据行号, 合计行号)
    """
    from openpyxl.styles import Font, PatternFill, Alignment

    # 表头
    headers = ['序号', '商品分类', '商品名称', '单位', '标准价', '折扣', '数量', '实际价', '小计']
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=h)
        _xl_header_style(cell)
    ws.row_dimensions[start_row].height = 22

    data_start = start_row + 1
    current_row = data_start

    for idx, item in enumerate(items, 1):
        std_price = item.get('标准价', 0)
        discount = item.get('折扣', 0)
        qty = item.get('数量', 1)
        is_gift = (std_price == '赠送' or std_price is None)

        # 交替行背景
        bg = 'F5F7FA' if idx % 2 == 0 else None

        # A: 序号
        c = ws.cell(row=current_row, column=1, value=idx)
        _xl_data_style(c, align='center', bg=bg)

        # B: 商品分类
        c = ws.cell(row=current_row, column=2, value=item.get('商品分类', ''))
        _xl_data_style(c, align='left', bg=bg)

        # C: 商品名称
        c = ws.cell(row=current_row, column=3, value=item.get('商品名称', ''))
        _xl_data_style(c, align='left', bg=bg)

        # D: 单位
        c = ws.cell(row=current_row, column=4, value=item.get('单位', ''))
        _xl_data_style(c, align='center', bg=bg)

        # E: 标准价
        if is_gift:
            c = ws.cell(row=current_row, column=5, value='赠送')
            _xl_data_style(c, align='center', bg=bg)
        else:
            c = ws.cell(row=current_row, column=5, value=float(std_price))
            _xl_data_style(c, align='right', bg=bg, num_format='#,##0.00')

        # F: 折扣（存为小数，格式化为百分比）
        if is_gift or discount == 0:
            c = ws.cell(row=current_row, column=6, value='-')
            _xl_data_style(c, align='center', bg=bg)
        else:
            c = ws.cell(row=current_row, column=6, value=float(discount))
            _xl_data_style(c, align='center', bg=bg, num_format='0%')

        # G: 数量
        c = ws.cell(row=current_row, column=7, value=int(qty))
        _xl_data_style(c, align='center', bg=bg)

        # H: 实际价
        if is_gift:
            c = ws.cell(row=current_row, column=8, value='赠送')
            _xl_data_style(c, align='center', bg=bg)
        elif compute_values:
            actual = float(std_price) * (1 - float(discount)) if discount else float(std_price)
            c = ws.cell(row=current_row, column=8, value=round(actual, 2))
            _xl_data_style(c, align='right', bg=bg, num_format='#,##0.00')
        elif discount == 0:
            c = ws.cell(row=current_row, column=8,
                        value=f'=E{current_row}')
            _xl_data_style(c, align='right', bg=bg, num_format='#,##0.00')
        else:
            c = ws.cell(row=current_row, column=8,
                        value=f'=E{current_row}*(1-F{current_row})')
            _xl_data_style(c, align='right', bg=bg, num_format='#,##0.00')

        # I: 小计
        if is_gift:
            c = ws.cell(row=current_row, column=9, value='赠送')
            _xl_data_style(c, align='center', bg=bg)
        elif compute_values:
            actual = float(std_price) * (1 - float(discount)) if discount else float(std_price)
            subtotal = round(actual * int(qty), 2)
            c = ws.cell(row=current_row, column=9, value=subtotal)
            _xl_data_style(c, align='right', bg=bg, num_format='#,##0.00')
        else:
            c = ws.cell(row=current_row, column=9,
                        value=f'=H{current_row}*G{current_row}')
            _xl_data_style(c, align='right', bg=bg, num_format='#,##0.00')

        ws.row_dimensions[current_row].height = 18
        current_row += 1

    last_data_row = current_row - 1

    # 合计行
    total_row = current_row
    ws.merge_cells(start_row=total_row, start_column=1,
                   end_row=total_row, end_column=6)
    c = ws.cell(row=total_row, column=1, value='合计')
    _xl_total_style(c, align='center')

    ws.cell(row=total_row, column=7, value='')
    _xl_total_style(ws.cell(row=total_row, column=7))

    ws.cell(row=total_row, column=8, value='')
    _xl_total_style(ws.cell(row=total_row, column=8))

    # 合计公式：SUM忽略文本（赠送）
    c = ws.cell(row=total_row, column=9,
                value=f'=SUM(I{data_start}:I{last_data_row})')
    _xl_total_style(c, align='right')
    c.number_format = '#,##0.00'

    ws.row_dimensions[total_row].height = 20

    # 应用边框（含表头）
    _xl_apply_border(ws, start_row, 1, total_row, 9)

    return last_data_row, total_row


def _xl_set_col_widths(ws):
    """设置标准列宽"""
    widths = {
        'A': 7,   # 序号
        'B': 16,  # 商品分类
        'C': 24,  # 商品名称
        'D': 10,  # 单位
        'E': 14,  # 标准价
        'F': 9,   # 折扣
        'G': 8,   # 数量
        'H': 14,  # 实际价
        'I': 16,  # 小计
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


# ============================================================
# 标准模板（≤50店）— Excel
# ============================================================
def generate_xlsx_standard(data, output_path):
    """生成标准单页 Excel 报价单（≤50店）"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = '报价单'
    ws.sheet_view.showGridLines = False

    _xl_set_col_widths(ws)

    # ── 标题区（行1-2） ──
    ws.merge_cells('A1:I1')
    c = ws.cell(row=1, column=1, value='"全来店"产品报价单')
    _xl_title_style(c, size=16)
    ws.row_dimensions[1].height = 36

    ws.merge_cells('A2:I2')
    c = ws.cell(row=2, column=1, value='上海收钱吧互联网科技股份有限公司')
    _xl_subtitle_style(c)
    ws.row_dimensions[2].height = 24

    # ── 空行 ──
    ws.row_dimensions[3].height = 6

    # ── 客户信息区（行4-7） ──
    client = data.get('客户信息', {})
    quote_no = data.get('报价编号', gen_quote_number())
    quote_date = data.get('报价日期', datetime.now().strftime('%Y年%m月%d日'))
    validity = data.get('报价有效期', '30个工作日')

    info_rows = [
        (f'致：{client.get("公司名称", "")}',   f'报价编号：{quote_no}'),
        (f'联系人：{client.get("联系人", "")}',  f'报价日期：{quote_date}'),
        (f'地址：{client.get("地址", "")}',       f'有效期：{validity}'),
        (f'电话：{client.get("电话", "")}',       ''),
    ]

    for i, (left, right) in enumerate(info_rows):
        row_num = 4 + i
        ws.merge_cells(start_row=row_num, start_column=1,
                       end_row=row_num, end_column=5)
        c = ws.cell(row=row_num, column=1, value=left)
        _xl_info_value_style(c)

        ws.merge_cells(start_row=row_num, start_column=6,
                       end_row=row_num, end_column=9)
        c = ws.cell(row=row_num, column=6, value=right)
        _xl_info_value_style(c)
        ws.row_dimensions[row_num].height = 18

    # ── 空行 ──
    ws.row_dimensions[8].height = 8

    # ── 报价明细表（从行9开始） ──
    items = data.get('报价项目', [])
    header_row = 9
    _, total_row = _xl_write_item_table(ws, items, header_row)

    # ── 金额大写（合计行下方） ──
    notes_row = total_row + 2
    total_cell_ref = f'I{total_row}'
    ws.merge_cells(start_row=notes_row, start_column=1,
                   end_row=notes_row, end_column=9)
    # 大写金额用Python计算（Excel没有内置大写金额函数）
    total_val = Decimal('0')
    for item in items:
        sp = item.get('标准价', 0)
        disc = item.get('折扣', 0)
        qty = item.get('数量', 1)
        if sp != '赠送' and sp is not None:
            ap = (Decimal(str(sp)) * (1 - Decimal(str(disc)))).quantize(Decimal('0.00'))
            total_val += ap * Decimal(str(qty))

    chinese_amt = number_to_chinese(float(total_val))
    c = ws.cell(row=notes_row, column=1,
                value=f'合计金额（大写）：{chinese_amt}')
    from openpyxl.styles import Font, Alignment
    c.font = Font(name='微软雅黑', size=10, bold=True)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[notes_row].height = 20

    # ── 备注条款 ──
    terms = data.get('条款', [
        '以上报价金额均为含税金额，税率为6%；',
        '报价有效期为30个工作日，自报价单生成之日起；',
        '具体折扣金额按签订合同（或销售订单）时具体数量确定价格；',
        '涉及短信、小程序授权、外卖平台接口调用等第三方机构收费部分，需单独计费；',
        '如需要三方代仓对接，需要一事一议。',
    ])

    terms_start = notes_row + 1
    ws.merge_cells(start_row=terms_start, start_column=1,
                   end_row=terms_start, end_column=9)
    c = ws.cell(row=terms_start, column=1, value='备注：')
    c.font = Font(name='微软雅黑', size=10, bold=True, color='CC8800')
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[terms_start].height = 18

    cn_nums = '①②③④⑤⑥⑦⑧⑨⑩'
    for i, term in enumerate(terms):
        r = terms_start + 1 + i
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
        prefix = cn_nums[i] if i < 10 else f'{i+1}.'
        c = ws.cell(row=r, column=1, value=f'{prefix} {term}')
        c.font = Font(name='微软雅黑', size=9, color='555555')
        c.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[r].height = 16

    # ── 页脚 ──
    footer_row = terms_start + 1 + len(terms) + 1
    ws.merge_cells(start_row=footer_row, start_column=1,
                   end_row=footer_row, end_column=9)
    c = ws.cell(row=footer_row, column=1,
                value='上海收钱吧互联网科技股份有限公司  |  地址：上海市闵行区浦江智慧广场陈行公路2168号7号楼')
    c.font = Font(name='微软雅黑', size=8, color='888888')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[footer_row].height = 16

    # 冻结表头
    ws.freeze_panes = f'A{header_row + 1}'

    wb.save(output_path)
    print(f'✅ Excel 报价单已生成：{output_path}')


# ============================================================
# 定制多页模板（>50店）— Excel
# ============================================================
def generate_xlsx_custom(data, output_path):
    """生成定制多 Sheet Excel 报价单（>50店）"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    # 删除默认 Sheet
    wb.remove(wb.active)

    client = data.get('客户信息', {})
    quote_no = data.get('报价编号', gen_quote_number())
    quote_date = data.get('报价日期', datetime.now().strftime('%Y年%m月%d日'))
    validity = data.get('报价有效期', '30个工作日')
    items = data.get('报价项目', [])

    # ── Sheet 1：封面 ──
    ws_cover = wb.create_sheet('封面')
    ws_cover.sheet_view.showGridLines = False
    ws_cover.column_dimensions['A'].width = 18
    ws_cover.column_dimensions['B'].width = 30

    ws_cover.row_dimensions[1].height = 20
    ws_cover.merge_cells('A2:B2')
    c = ws_cover.cell(row=2, column=1, value='"全来店"产品报价方案')
    _xl_title_style(c, size=18)
    ws_cover.row_dimensions[2].height = 44

    ws_cover.merge_cells('A3:B3')
    c = ws_cover.cell(row=3, column=1, value='上海收钱吧互联网科技股份有限公司')
    _xl_subtitle_style(c)
    ws_cover.row_dimensions[3].height = 28

    ws_cover.row_dimensions[4].height = 12

    cover_info = [
        ('客户名称', client.get('公司名称', '')),
        ('联系人',   client.get('联系人', '')),
        ('联系电话', client.get('电话', '')),
        ('报价编号', quote_no),
        ('报价日期', quote_date),
        ('有效期',   validity),
    ]

    for i, (label, value) in enumerate(cover_info):
        r = 5 + i
        c_label = ws_cover.cell(row=r, column=1, value=label)
        _xl_info_label_style(c_label)
        c_val = ws_cover.cell(row=r, column=2, value=value)
        _xl_info_value_style(c_val)
        ws_cover.row_dimensions[r].height = 22

    _xl_apply_border(ws_cover, 5, 1, 5 + len(cover_info) - 1, 2)

    # 记录封面下一可用行（用于后续追加汇总内容）
    cover_summary_start_row = 5 + len(cover_info) + 2

    # ── 按模块分类（不含硬件设备）──
    categories = {
        '门店软件套餐': [],
        '门店增值模块': [],
        '总部模块': [],
        '实施服务': [],
    }

    for item in items:
        cat = item.get('模块分类', '门店软件套餐')
        if cat in categories:
            categories[cat].append(item)
        elif cat != '硬件设备':
            categories['门店软件套餐'].append(item)

    # ── 各分类 Sheet ──
    # 门店软件套餐 + 门店增值模块 合并为一个 Sheet；硬件设备不纳入报价
    # 定价规则：软件/总部 → qty=1, 折扣20%；实施服务 → qty=1, 无折扣

    def _override_items(src_items, discount, qty=1):
        """返回 qty/discount 覆盖后的副本（刊例价展示用）"""
        result = []
        for it in src_items:
            ni = dict(it)
            ni['数量'] = qty
            if ni.get('标准价') not in ('赠送', None):
                ni['折扣'] = discount
            result.append(ni)
        return result

    MERGED_SHEET_CATS = ['门店软件套餐', '门店增值模块']
    cat_totals = {}  # {cat_name: (total_cell_ref, sheet_title)}

    # ── 合并 Sheet：门店软件与增值（qty=1, 折扣20%）──
    merged_has_items = any(categories[c] for c in MERGED_SHEET_CATS)
    if merged_has_items:
        ws_merged = wb.create_sheet('门店软件与增值')
        ws_merged.sheet_view.showGridLines = False
        _xl_set_col_widths(ws_merged)

        ws_merged.merge_cells('A1:I1')
        c = ws_merged.cell(row=1, column=1, value='门店软件与增值模块')
        _xl_title_style(c, size=14)
        ws_merged.row_dimensions[1].height = 30
        ws_merged.row_dimensions[2].height = 8

        section_row = 3
        for cat_name in MERGED_SHEET_CATS:
            cat_items = categories[cat_name]
            if not cat_items:
                continue
            # 分区标题行
            ws_merged.merge_cells(start_row=section_row, start_column=1,
                                  end_row=section_row, end_column=9)
            c = ws_merged.cell(row=section_row, column=1, value=cat_name)
            c.font = Font(name='微软雅黑', bold=True, size=10, color='CC8800')
            c.fill = PatternFill('solid', fgColor='FFFBF0')
            c.alignment = Alignment(horizontal='left', vertical='center')
            ws_merged.row_dimensions[section_row].height = 18
            section_row += 1

            display_items = _override_items(cat_items, discount=0.2, qty=1)
            _, total_row = _xl_write_item_table(ws_merged, display_items, section_row, compute_values=True)
            cat_totals[cat_name] = (f'I{total_row}', ws_merged.title)
            section_row = total_row + 2

        ws_merged.freeze_panes = 'A4'

    # ── 总部模块 Sheet（qty=1, 折扣20%）──
    if categories.get('总部模块'):
        ws = wb.create_sheet('总部模块')
        ws.sheet_view.showGridLines = False
        _xl_set_col_widths(ws)
        ws.merge_cells('A1:I1')
        c = ws.cell(row=1, column=1, value='总部模块')
        _xl_title_style(c, size=14)
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 8
        display_items = _override_items(categories['总部模块'], discount=0.2, qty=1)
        _, total_row = _xl_write_item_table(ws, display_items, 3, compute_values=True)
        ws.freeze_panes = 'A4'
        cat_totals['总部模块'] = (f'I{total_row}', ws.title)

    # ── 实施服务 Sheet（qty=1, 无折扣）──
    if categories.get('实施服务'):
        ws = wb.create_sheet('实施服务')
        ws.sheet_view.showGridLines = False
        _xl_set_col_widths(ws)
        ws.merge_cells('A1:I1')
        c = ws.cell(row=1, column=1, value='实施服务')
        _xl_title_style(c, size=14)
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 8
        display_items = _override_items(categories['实施服务'], discount=0, qty=1)
        _, total_row = _xl_write_item_table(ws, display_items, 3, compute_values=True)
        ws.freeze_panes = 'A4'
        cat_totals['实施服务'] = (f'I{total_row}', ws.title)

    # ── 封面追加：条款说明 ──
    r = cover_summary_start_row

    # 条款
    terms = data.get('条款', [
        '以上报价金额均为含税金额，税率为6%；',
        '报价有效期为30个工作日，自报价单生成之日起；',
        '具体折扣金额按签订合同（或销售订单）时具体数量确定价格；',
        '涉及短信、小程序授权、外卖平台接口调用等第三方机构收费部分，需单独计费；',
        '如需要三方代仓对接，需要一事一议。',
    ])
    ws_cover.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    c = ws_cover.cell(row=r, column=1, value='备注与条款：')
    c.font = Font(name='微软雅黑', size=10, bold=True, color='CC8800')
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws_cover.row_dimensions[r].height = 18
    r += 1

    cn_nums = '①②③④⑤⑥⑦⑧⑨⑩'
    for i, term in enumerate(terms):
        ws_cover.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        prefix = cn_nums[i] if i < 10 else f'{i+1}.'
        c = ws_cover.cell(row=r, column=1, value=f'{prefix} {term}')
        c.font = Font(name='微软雅黑', size=9, color='555555')
        c.alignment = Alignment(horizontal='left', vertical='center')
        ws_cover.row_dimensions[r].height = 16
        r += 1

    # 阶梯报价参考 Sheet（如有配置）
    _xl_add_tiered_sheet(wb, data)

    wb.save(output_path)
    print(f'✅ Excel 报价单已生成：{output_path}')


# ============================================================
# 阶梯报价参考 Sheet（Excel，定制版附页）
# ============================================================
def _xl_add_tiered_sheet(wb, data):
    """在 Excel 工作簿末尾添加阶梯报价参考 Sheet"""
    tiers = data.get('阶梯配置', [])
    if not tiers:
        return

    from openpyxl.styles import Font, PatternFill, Alignment

    ws = wb.create_sheet('阶梯报价参考')
    ws.sheet_view.showGridLines = False

    items = data.get('报价项目', [])
    n_tiers = len(tiers)

    # 列宽
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 14
    tier_col_letters = ['D', 'E', 'F', 'G', 'H'][:n_tiers]
    for col in tier_col_letters:
        ws.column_dimensions[col].width = 18

    last_col = tier_col_letters[-1]

    # 标题
    ws.merge_cells(f'A1:{last_col}1')
    c = ws.cell(row=1, column=1, value='阶梯报价参考')
    _xl_title_style(c, size=14)
    ws.row_dimensions[1].height = 32

    # 表头
    def tier_label(t):
        d = t.get('折扣', 0)
        if d == 0:
            return f"{t['标签']}（标准价）"
        zhe = round((1 - d) * 10, 1)
        zhe_str = f"{int(zhe)}折" if zhe == int(zhe) else f"{zhe}折"
        return f"{t['标签']}（{zhe_str}）"

    headers = ['商品名称', '单位', '标准价'] + [tier_label(t) for t in tiers]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        _xl_header_style(c)
    ws.row_dimensions[2].height = 22

    current_row = 3

    cat_order = ['门店软件套餐', '门店增值模块', '总部模块', '实施服务']
    categories = {k: [] for k in cat_order}
    for item in items:
        cat = item.get('模块分类', '门店软件套餐')
        if cat == '硬件设备':
            continue
        if cat in categories:
            categories[cat].append(item)
        else:
            categories['门店软件套餐'].append(item)

    tier_grand_totals = [Decimal('0')] * n_tiers
    NO_TIER_DISCOUNT_CATS = {'实施服务'}

    for cat_name in cat_order:
        cat_items = categories[cat_name]
        if not cat_items:
            continue

        # 分类标题行
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=3 + n_tiers)
        c = ws.cell(row=current_row, column=1, value=cat_name)
        c.font = Font(name='微软雅黑', bold=True, size=10, color='CC8800')
        c.fill = PatternFill('solid', fgColor='FFFBF0')
        c.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        cat_tier_totals = [Decimal('0')] * n_tiers
        apply_tier_discount = cat_name not in NO_TIER_DISCOUNT_CATS

        for item in cat_items:
            std_price = item.get('标准价', 0)
            unit = item.get('单位', '')
            item_qty = item.get('数量', 1)
            is_per_store = '店' in unit

            c = ws.cell(row=current_row, column=1, value=item.get('商品名称', ''))
            c.font = Font(name='微软雅黑', size=9)
            c.alignment = Alignment(horizontal='left', vertical='center')

            c = ws.cell(row=current_row, column=2, value=unit)
            c.font = Font(name='微软雅黑', size=9)
            c.alignment = Alignment(horizontal='center', vertical='center')

            if std_price == '赠送' or std_price is None:
                c = ws.cell(row=current_row, column=3, value='赠送')
                c.font = Font(name='微软雅黑', size=9)
                c.alignment = Alignment(horizontal='center', vertical='center')
                for ci in range(n_tiers):
                    c = ws.cell(row=current_row, column=4 + ci, value='赠送')
                    c.font = Font(name='微软雅黑', size=9)
                    c.alignment = Alignment(horizontal='center', vertical='center')
            else:
                std_d = Decimal(str(std_price))
                c = ws.cell(row=current_row, column=3, value=float(std_d))
                c.font = Font(name='微软雅黑', size=9)
                c.number_format = '#,##0.00'
                c.alignment = Alignment(horizontal='right', vertical='center')
                for ti, t in enumerate(tiers):
                    d = Decimal(str(t.get('折扣', 0))) if apply_tier_discount else Decimal('0')
                    qty = Decimal(str(t['门店数'])) if is_per_store else Decimal(str(item_qty))
                    actual = (std_d * (1 - d)).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)
                    subtotal = (actual * qty).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)
                    cat_tier_totals[ti] += subtotal
                    c = ws.cell(row=current_row, column=4 + ti, value=float(subtotal))
                    c.font = Font(name='微软雅黑', size=9)
                    c.number_format = '#,##0.00'
                    c.alignment = Alignment(horizontal='right', vertical='center')

            ws.row_dimensions[current_row].height = 18
            current_row += 1

        # 分类小计行
        c = ws.cell(row=current_row, column=3, value='小计')
        c.font = Font(name='微软雅黑', bold=True, size=9)
        c.fill = PatternFill('solid', fgColor='FFF5D6')
        c.alignment = Alignment(horizontal='center', vertical='center')
        for ti, tot in enumerate(cat_tier_totals):
            tier_grand_totals[ti] += tot
            c = ws.cell(row=current_row, column=4 + ti, value=float(tot))
            c.font = Font(name='微软雅黑', bold=True, size=9)
            c.number_format = '#,##0.00'
            c.fill = PatternFill('solid', fgColor='FFF5D6')
            c.alignment = Alignment(horizontal='right', vertical='center')
        ws.row_dimensions[current_row].height = 18
        current_row += 1

    # 合计行
    c = ws.cell(row=current_row, column=3, value='合计')
    c.font = Font(name='微软雅黑', bold=True, size=10, color='CC8800')
    c.fill = PatternFill('solid', fgColor='FFE082')
    c.alignment = Alignment(horizontal='center', vertical='center')
    for ti, tot in enumerate(tier_grand_totals):
        c = ws.cell(row=current_row, column=4 + ti, value=float(tot))
        c.font = Font(name='微软雅黑', bold=True, size=10, color='CC8800')
        c.number_format = '#,##0.00'
        c.fill = PatternFill('solid', fgColor='FFE082')
        c.alignment = Alignment(horizontal='right', vertical='center')
    ws.row_dimensions[current_row].height = 22
    grand_total_row = current_row
    current_row += 1

    # 折算单店年费行
    ws.merge_cells(start_row=current_row, start_column=1,
                   end_row=current_row, end_column=3)
    c = ws.cell(row=current_row, column=1, value='折算单店年费')
    c.font = Font(name='微软雅黑', bold=True, size=9)
    c.fill = PatternFill('solid', fgColor='FFF5D6')
    c.alignment = Alignment(horizontal='left', vertical='center')
    for ti, t in enumerate(tiers):
        stores = Decimal(str(t['门店数']))
        per_store = (tier_grand_totals[ti] / stores).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)
        c = ws.cell(row=current_row, column=4 + ti, value=float(per_store))
        c.font = Font(name='微软雅黑', size=9)
        c.number_format = '#,##0.00'
        c.fill = PatternFill('solid', fgColor='FFF5D6')
        c.alignment = Alignment(horizontal='right', vertical='center')
    ws.row_dimensions[current_row].height = 18

    _xl_apply_border(ws, 2, 1, current_row, 3 + n_tiers)


# ============================================================
# 利润测算（终端输出，不写入文档）
# ============================================================
def calc_profit(data, cost_data):
    """计算利润并在终端输出"""
    items = data.get('报价项目', [])

    print("\n" + "="*70)
    print("  利润测算报告（内部参考，严禁外泄）")
    print("="*70)
    print(f"{'商品名称':<20} {'底价':>8} {'售价':>8} {'数量':>4} {'利润':>10} {'利润率':>8}")
    print("-"*70)

    total_cost = Decimal('0')
    total_revenue = Decimal('0')

    for item in items:
        name = item.get('商品名称', '')
        std_price = item.get('标准价', 0)
        discount = item.get('折扣', 0)
        qty = item.get('数量', 1)

        if std_price == '赠送' or std_price is None:
            print(f"{name:<20} {'赠送':>8} {'赠送':>8} {qty:>4} {'-':>10} {'-':>8}")
            continue

        std_price_d = Decimal(str(std_price))
        discount_d = Decimal(str(discount))
        qty_d = Decimal(str(qty))
        actual_price = (std_price_d * (1 - discount_d)).quantize(Decimal('0.00'))

        # 查找底价
        cost_key = name
        cost = cost_data.get(cost_key, None)

        if cost is None or cost == '赠送':
            print(f"{name:<20} {'未知':>8} {float(actual_price):>8.0f} {qty:>4} {'未知':>10} {'未知':>8}")
            total_revenue += actual_price * qty_d
            continue

        cost_d = Decimal(str(cost))
        item_profit = (actual_price - cost_d) * qty_d
        item_revenue = actual_price * qty_d

        if actual_price > 0:
            margin = float((actual_price - cost_d) / actual_price * 100)
        else:
            margin = 0

        total_cost += cost_d * qty_d
        total_revenue += item_revenue

        # 警告：售价低于底价
        warning = " ⚠️ 低于底价!" if actual_price < cost_d else ""

        print(f"{name:<20} {float(cost_d):>8.0f} {float(actual_price):>8.0f} {qty:>4} {float(item_profit):>10,.0f} {margin:>7.1f}%{warning}")

    total_profit = total_revenue - total_cost
    overall_margin = float(total_profit / total_revenue * 100) if total_revenue > 0 else 0

    print("-"*70)
    print(f"{'合计':<20} {float(total_cost):>8,.0f} {float(total_revenue):>8,.0f} {'':>4} {float(total_profit):>10,.0f} {overall_margin:>7.1f}%")
    print("="*70)
    print(f"  年化总收入：¥{float(total_revenue):,.2f}")
    print(f"  年化总成本：¥{float(total_cost):,.2f}")
    print(f"  年化总利润：¥{float(total_profit):,.2f}")
    print(f"  整体利润率：{overall_margin:.1f}%")
    print("="*70)

# ============================================================
# 主入口
# ============================================================
def main():
    parser = argparse.ArgumentParser(description='全来店报价单 PDF + Excel 生成器')
    parser.add_argument('--config', required=True, help='JSON配置文件路径')
    parser.add_argument('--output', required=True, help='输出PDF路径')
    parser.add_argument('--output-xlsx', help='输出Excel路径（可选，默认与PDF同名.xlsx）')
    parser.add_argument('--no-xlsx', action='store_true', help='跳过Excel生成，仅生成PDF')
    parser.add_argument('--profit', action='store_true', help='输出利润测算')
    parser.add_argument('--cost-data', help='底价JSON文件路径（利润测算用）')
    args = parser.parse_args()

    # 读取配置
    with open(args.config, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # 判断模板类型
    store_count = data.get('门店数量', 1)
    use_custom = store_count > 50

    # ── 生成 PDF ──
    styles = get_styles()
    if use_custom:
        story = build_custom_template(data, styles)
    else:
        story = build_standard_template(data, styles)

    doc = SimpleDocTemplate(
        args.output,
        pagesize=A4,
        topMargin=15*mm,
        bottomMargin=15*mm,
        leftMargin=15*mm,
        rightMargin=15*mm,
    )
    doc.build(story)
    print(f'\n✅ PDF 报价单已生成：{args.output}')

    # ── 生成 Excel ──
    if not args.no_xlsx:
        # 默认 xlsx 路径：与 PDF 同名，后缀改为 .xlsx
        if args.output_xlsx:
            xlsx_path = args.output_xlsx
        else:
            base = os.path.splitext(args.output)[0]
            xlsx_path = base + '.xlsx'

        if use_custom:
            generate_xlsx_custom(data, xlsx_path)
        else:
            generate_xlsx_standard(data, xlsx_path)

    # ── 利润测算 ──
    if args.profit and args.cost_data:
        with open(args.cost_data, 'r', encoding='utf-8') as f:
            cost_data = json.load(f)
        calc_profit(data, cost_data)
    elif args.profit:
        print("\n⚠️ 需要提供 --cost-data 参数才能进行利润测算")

if __name__ == '__main__':
    main()
