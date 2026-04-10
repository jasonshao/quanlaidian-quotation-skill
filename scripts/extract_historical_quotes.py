#!/usr/bin/env python3
"""Extract normalized quote cases from historical Excel files."""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Iterator
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


def iter_excel_files(history_root: Path) -> Iterator[Path]:
    for path in sorted(history_root.rglob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        yield path


def infer_brand_name(path: Path) -> str:
    stem = path.stem
    for token in ("报价单", "对客报价", "报价", "确认单", "供应链"):
        stem = stem.replace(token, "")
    return stem.strip("-_（）() ") or path.parent.name


def store_count_band(store_count: int | None) -> str | None:
    if store_count is None:
        return None
    if store_count <= 10:
        return "1-10"
    if store_count <= 50:
        return "11-50"
    if store_count <= 300:
        return "51-300"
    return "300+"


def infer_meal_type(sheet_title: str, path: Path) -> str | None:
    haystack = f"{sheet_title} {path.name}"
    if "轻餐" in haystack:
        return "轻餐"
    if "正餐" in haystack:
        return "正餐"
    return None


def infer_quote_kind(sheet_title: str, path: Path) -> str | None:
    haystack = f"{sheet_title} {path.name}"
    if "对外" in haystack or "对客" in haystack:
        return "对外报价"
    if "底价" in haystack:
        return "底价单"
    if "确认" in haystack:
        return "确认单"
    return None


def normalize_number(value):
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if text == "赠送":
        return None
    try:
        return float(text)
    except ValueError:
        return None


def build_case_stub(path: Path) -> dict:
    return {
        "case_id": path.stem,
        "source_file": str(path),
        "brand_name": infer_brand_name(path),
        "meal_type": None,
        "quote_kind": None,
        "store_count": None,
        "store_count_band": None,
        "selected_package": None,
        "store_modules": [],
        "hq_modules": [],
        "implementation_service": None,
        "discounted_total": None,
        "line_items": [],
        "notes": [],
        "raw_extract_status": "stub_only",
    }


def build_open_failed_case(path: Path, error: Exception) -> dict:
    case = build_case_stub(path)
    case["raw_extract_status"] = "open_failed"
    case["notes"].append(f"open_failed: {type(error).__name__}: {error}")
    return case


def find_header_row(ws) -> int | None:
    for row_idx in range(1, min(ws.max_row, 20) + 1):
        row_values = [ws.cell(row=row_idx, column=col).value for col in range(1, 13)]
        text = " ".join(str(v) for v in row_values if v not in (None, ""))
        if "商品名称" in text and ("总价" in text or "金额" in text):
            return row_idx
    return None


def build_column_map(ws, header_row: int) -> dict[str, int]:
    aliases = {
        "module": ("模块",),
        "category": ("商品分类",),
        "product_name": ("商品名称",),
        "unit": ("单位",),
        "price": ("折扣后单价", "标准价"),
        "quantity": ("购买数量", "购买数量\n(不需要的功能请空着)"),
        "subtotal": ("总价", "金额"),
        "note": ("套餐说明/其他说明", "功能说明", "备注"),
    }
    header_map: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=header_row, column=col).value
        if value in (None, ""):
            continue
        label = str(value).strip()
        for key, options in aliases.items():
            if label in options and key not in header_map:
                header_map[key] = col
    return header_map


def read_mapped_value(ws, row_idx: int, header_map: dict[str, int], key: str):
    column = header_map.get(key)
    if column is None:
        return None
    return ws.cell(row=row_idx, column=column).value


def extract_case_from_workbook(path: Path) -> dict:
    wb = load_workbook(path, data_only=True)
    case = build_case_stub(path)

    for ws in wb.worksheets:
        header_row = find_header_row(ws)
        if header_row is None:
            continue
        header_map = build_column_map(ws, header_row)

        case["meal_type"] = case["meal_type"] or infer_meal_type(ws.title, path)
        case["quote_kind"] = case["quote_kind"] or infer_quote_kind(ws.title, path)

        for row_idx in range(header_row + 1, ws.max_row + 1):
            module = read_mapped_value(ws, row_idx, header_map, "module")
            category = read_mapped_value(ws, row_idx, header_map, "category")
            product_name = read_mapped_value(ws, row_idx, header_map, "product_name")
            unit = read_mapped_value(ws, row_idx, header_map, "unit")
            price = read_mapped_value(ws, row_idx, header_map, "price")
            quantity = read_mapped_value(ws, row_idx, header_map, "quantity")
            subtotal = read_mapped_value(ws, row_idx, header_map, "subtotal")
            note = read_mapped_value(ws, row_idx, header_map, "note")

            if module and not product_name:
                note_parts = [
                    str(value).strip()
                    for value in (module, category, unit, note)
                    if value not in (None, "")
                ]
                if note_parts:
                    case["notes"].append(" | ".join(note_parts))
                if category not in (None, ""):
                    case["notes"].append(str(category).strip())
                if note not in (None, ""):
                    case["notes"].append(str(note).strip())
                continue

            if note not in (None, ""):
                case["notes"].append(str(note).strip())

            if not product_name:
                continue

            product_name = str(product_name).strip()
            module_text = str(module).strip() if module not in (None, "") else ""
            category_text = str(category).strip() if category not in (None, "") else ""
            unit_text = str(unit).strip() if unit not in (None, "") else None
            section_text = module_text or category_text

            if "硬件" in module_text or "硬件" in category_text or "硬件" in product_name:
                continue

            line_item = {
                "category": section_text or None,
                "product_name": product_name,
                "unit": unit_text,
                "discounted_unit_price": normalize_number(price),
                "quantity": normalize_number(quantity),
                "subtotal": normalize_number(subtotal),
            }
            case["line_items"].append(line_item)

            if (
                "门店优惠套餐" in section_text or category_text == "套餐"
            ) and case["selected_package"] is None:
                case["selected_package"] = product_name
                if line_item["quantity"] is not None:
                    case["store_count"] = int(line_item["quantity"])

            if "门店增值模块" in section_text and product_name not in case["store_modules"]:
                case["store_modules"].append(product_name)

            if "总部模块" in section_text and product_name not in case["hq_modules"]:
                case["hq_modules"].append(product_name)

            if (
                "实施服务" in section_text
                or "实施与售后服务" in category_text
            ) and case["implementation_service"] is None:
                case["implementation_service"] = product_name

    case["store_count_band"] = store_count_band(case["store_count"])
    case["discounted_total"] = round(
        sum(item["subtotal"] or 0 for item in case["line_items"]),
        2,
    ) or None
    case["raw_extract_status"] = "parsed" if case["line_items"] else "stub_only"
    return case


def extract_history_casebase(history_root: Path) -> list[dict]:
    cases: list[dict] = []
    for path in iter_excel_files(history_root):
        try:
            cases.append(extract_case_from_workbook(path))
        except (BadZipFile, OSError, ValueError, InvalidFileException, KeyError) as error:
            cases.append(build_open_failed_case(path, error))
    return cases


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Extract historical quote casebase")
    parser.add_argument("--history-root", required=True, help="Root directory of historical quotes")
    parser.add_argument("--output", required=True, help="JSONL output path")
    args = parser.parse_args(argv)

    history_root = Path(args.history_root).resolve()
    cases = extract_history_casebase(history_root)
    output_path = Path(args.output).resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as fh:
        for item in cases:
            fh.write(json.dumps(item, ensure_ascii=False) + "\n")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
